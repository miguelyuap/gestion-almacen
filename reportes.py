"""
Módulo de Generación de Reportes Empresariales (Excel & PDF) para gestion-almacen.
- Excel: openpyxl (Archivos masivos .xlsx estilizados, formato moneda COP, auto-ancho de columnas).
- PDF: reportlab (Documentos formales .pdf: Cierre Z de Caja y Comprobante de Venta).
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import (
    HRFlowable,
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


# ---------------------------------------------------------------------------
# ESTILOS DE EXCEL (openpyxl)
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=15, bold=True, color="0F172A")
SUBTITLE_FONT = Font(name="Arial", size=9, italic=True, color="64748B")
DATA_FONT = Font(name="Arial", size=10, color="334155")
TOTAL_FONT = Font(name="Arial", size=11, bold=True, color="0F172A")

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

TOTAL_BORDER = Border(
    top=Side(style="thin", color="0F172A"),
    bottom=Side(style="double", color="0F172A"),
)

CURRENCY_FORMAT = '"$"#,##0.00'
NUMBER_FORMAT = "#,##0"


def _auto_ajustar_columnas(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Ajusta automáticamente el ancho de las columnas de la hoja de Excel según el contenido."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if cell.number_format == CURRENCY_FORMAT:
                val_str = f"${val_str}"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


# ---------------------------------------------------------------------------
# REPORTES EN EXCEL
# ---------------------------------------------------------------------------

def generar_excel_inventario(productos: list[dict[str, Any]]) -> io.BytesIO:
    """Genera reporte de inventario actual y valoración en Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Valorizado"

    # Encabezado del reporte
    ws.merge_cells("A1:F1")
    ws["A1"] = "ERP ALMACÉN — REPORTE DE INVENTARIO VALORIZADO"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = SUBTITLE_FONT

    ws.append([])  # Fila vacía

    # Cabeceras de la tabla
    headers = ["Código (SKU)", "Nombre del Producto", "Precio Venta", "Stock Total", "Stock Mínimo", "Valor Inventario ($)"]
    ws.append(headers)
    
    header_row_idx = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Filas de datos
    start_data_row = 5
    for p_raw in productos:
        p = dict(p_raw) if hasattr(p_raw, "keys") else p_raw
        stock = p.get("stock_total", 0)
        precio = float(p.get("precio_venta", 0))
        valor_inv = stock * precio

        row = [
            p.get("codigo", ""),
            p.get("nombre", ""),
            precio,
            stock,
            p.get("stock_minimo", 5),
            valor_inv,
        ]
        ws.append(row)

        curr_row = ws.max_row
        ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=3).number_format = CURRENCY_FORMAT
        ws.cell(row=curr_row, column=4).number_format = NUMBER_FORMAT
        ws.cell(row=curr_row, column=5).number_format = NUMBER_FORMAT
        ws.cell(row=curr_row, column=6).number_format = CURRENCY_FORMAT

        for col_idx in range(1, 7):
            c = ws.cell(row=curr_row, column=col_idx)
            c.font = DATA_FONT
            c.border = THIN_BORDER

    # Fila de Totales
    total_row_idx = ws.max_row + 1
    ws.cell(row=total_row_idx, column=2, value="TOTALES GENERALES").font = TOTAL_FONT
    ws.cell(row=total_row_idx, column=4, value=f"=SUM(D5:D{total_row_idx-1})").number_format = NUMBER_FORMAT
    ws.cell(row=total_row_idx, column=6, value=f"=SUM(F5:F{total_row_idx-1})").number_format = CURRENCY_FORMAT

    for col_idx in range(1, 7):
        c = ws.cell(row=total_row_idx, column=col_idx)
        c.font = TOTAL_FONT
        c.border = TOTAL_BORDER

    _auto_ajustar_columnas(ws)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def generar_excel_ventas(ventas: list[dict[str, Any]]) -> io.BytesIO:
    """Genera reporte histórico de ventas POS en Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas Realizadas"

    ws.merge_cells("A1:H1")
    ws["A1"] = "ERP ALMACÉN — HISTORIAL DE VENTAS POS"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = SUBTITLE_FONT

    ws.append([])

    headers = ["Nº Factura", "Fecha / Hora", "Cliente", "NIT / Cédula", "Método Pago", "Subtotal (Sin IVA)", "IVA (19%)", "Total Pagado"]
    ws.append(headers)

    header_row_idx = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for v_raw in ventas:
        v = dict(v_raw) if hasattr(v_raw, "keys") else v_raw
        if v.get("estado") == "anulada":
            continue

        total = float(v.get("total", 0))
        subtotal = float(v.get("subtotal", total / 1.19))
        iva = float(v.get("iva_total", total - subtotal))

        row = [
            v.get("numero_factura", ""),
            v.get("fecha", ""),
            v.get("cliente_nombre", ""),
            v.get("cliente_nit", ""),
            v.get("metodo_pago", ""),
            subtotal,
            iva,
            total,
        ]
        ws.append(row)

        curr_row = ws.max_row
        ws.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=6).number_format = CURRENCY_FORMAT
        ws.cell(row=curr_row, column=7).number_format = CURRENCY_FORMAT
        ws.cell(row=curr_row, column=8).number_format = CURRENCY_FORMAT

        for col_idx in range(1, 9):
            c = ws.cell(row=curr_row, column=col_idx)
            c.font = DATA_FONT
            c.border = THIN_BORDER

    # Totales
    total_row_idx = ws.max_row + 1
    ws.cell(row=total_row_idx, column=5, value="TOTAL VENDIDO").font = TOTAL_FONT
    ws.cell(row=total_row_idx, column=6, value=f"=SUM(F5:F{total_row_idx-1})").number_format = CURRENCY_FORMAT
    ws.cell(row=total_row_idx, column=7, value=f"=SUM(G5:G{total_row_idx-1})").number_format = CURRENCY_FORMAT
    ws.cell(row=total_row_idx, column=8, value=f"=SUM(H5:H{total_row_idx-1})").number_format = CURRENCY_FORMAT

    for col_idx in range(1, 9):
        c = ws.cell(row=total_row_idx, column=col_idx)
        c.font = TOTAL_FONT
        c.border = TOTAL_BORDER

    _auto_ajustar_columnas(ws)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def generar_excel_kardex(kardex_list: list[Any]) -> io.BytesIO:
    """Genera reporte completo de Kárdex en Excel."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos Kardex"

    ws.merge_cells("A1:G1")
    ws["A1"] = "ERP ALMACÉN — REGISTRO HISTÓRICO DE KÁRDEX"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = SUBTITLE_FONT

    ws.append([])

    headers = ["Fecha / Hora", "SKU", "Producto", "Bodega", "Movimiento", "Cantidad", "Costo Unitario ($)"]
    ws.append(headers)

    header_row_idx = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for k in kardex_list:
        k_dict = dict(k) if hasattr(k, "keys") else k
        row = [
            k_dict.get("fecha", ""),
            k_dict.get("producto_codigo", ""),
            k_dict.get("producto_nombre", ""),
            k_dict.get("almacen_nombre", ""),
            k_dict.get("tipo_movimiento", ""),
            k_dict.get("cantidad", 0),
            float(k_dict.get("costo_unitario", 0)),
        ]
        ws.append(row)

        curr_row = ws.max_row
        ws.cell(row=curr_row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=curr_row, column=6).number_format = NUMBER_FORMAT
        ws.cell(row=curr_row, column=7).number_format = CURRENCY_FORMAT

        for col_idx in range(1, 8):
            c = ws.cell(row=curr_row, column=col_idx)
            c.font = DATA_FONT
            c.border = THIN_BORDER

    _auto_ajustar_columnas(ws)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# ---------------------------------------------------------------------------
# REPORTES EN PDF (reportlab)
# ---------------------------------------------------------------------------

def generar_pdf_cierre_caja(resumen_dia: dict[str, Any], ventas_dia: list[dict[str, Any]], fecha_str: str) -> io.BytesIO:
    """Genera el reporte formal de Cierre Z de Caja del Día en PDF."""
    stream = io.BytesIO()
    doc = SimpleDocTemplate(
        stream,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PDFTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#0F172A"),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "PDFSubtitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        textColor=colors.HexColor("#64748B"),
        spaceAfter=12,
    )
    section_title = ParagraphStyle(
        "PDFSection",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=colors.HexColor("#0F172A"),
        spaceBefore=10,
        spaceAfter=6,
    )
    cell_style = ParagraphStyle("Cell", fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#334155"))
    cell_bold = ParagraphStyle("CellB", fontName="Helvetica-Bold", fontSize=9, textColor=colors.HexColor("#0F172A"))

    elements = []

    # Encabezado Corporativo
    elements.append(Paragraph("ERP ALMACÉN S.A.S.", title_style))
    elements.append(Paragraph(f"COMPROBANTE DE CIERRE Z DE CAJA — FECHA: {fecha_str}", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#0F172A"), spaceAfter=12))

    # Resumen General de Ventas
    elements.append(Paragraph("1. Resumen Consolidado de la Jornada", section_title))

    total_ventas = float(resumen_dia.get("total_ventas", 0))
    efectivo = float(resumen_dia.get("efectivo", 0))
    tarjeta = float(resumen_dia.get("tarjeta", 0))
    transferencia = float(resumen_dia.get("transferencia", 0))
    num_trans = int(resumen_dia.get("num_transacciones", 0))

    data_resumen = [
        [Paragraph("Concepto", cell_bold), Paragraph("Valor ($ COP)", cell_bold)],
        [Paragraph("Total Recaudado en Efectivo", cell_style), Paragraph(f"${efectivo:,.2f}", cell_bold)],
        [Paragraph("Total Tarjeta Débito/Crédito", cell_style), Paragraph(f"${tarjeta:,.2f}", cell_bold)],
        [Paragraph("Total Transferencia Bancaria", cell_style), Paragraph(f"${transferencia:,.2f}", cell_bold)],
        [Paragraph("Transacciones Totales Procesadas", cell_style), Paragraph(str(num_trans), cell_style)],
        [Paragraph("TOTAL GENERAL CAJA (Z)", cell_bold), Paragraph(f"${total_ventas:,.2f}", cell_bold)],
    ]

    t_resumen = Table(data_resumen, colWidths=[300, 240])
    t_resumen.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#F8FAFC")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("BACKGROUND", (0, -1), (1, -1), colors.HexColor("#EFF6FF")),
        ])
    )
    elements.append(t_resumen)
    elements.append(Spacer(1, 14))

    # Detalle de Transacciones del Día
    elements.append(Paragraph("2. Detalle de Facturas Emitidas", section_title))
    headers_tx = ["Nº Factura", "Cliente / NIT", "Método Pago", "Total Factura"]
    table_tx_data = [[Paragraph(h, cell_bold) for h in headers_tx]]

    for v_raw in ventas_dia:
        v = dict(v_raw) if hasattr(v_raw, "keys") else v_raw
        table_tx_data.append([
            Paragraph(v.get("numero_factura", ""), cell_style),
            Paragraph(f"{v.get('cliente_nombre', '')} ({v.get('cliente_nit', '')})", cell_style),
            Paragraph(v.get("metodo_pago", ""), cell_style),
            Paragraph(f"${float(v.get('total', 0)):,.2f}", cell_bold),
        ])

    t_tx = Table(table_tx_data, colWidths=[100, 220, 110, 110])
    t_tx.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ])
    )
    # Cambiar color de texto en cabecera PDF
    for i in range(len(headers_tx)):
        t_tx.setStyle(TableStyle([("TEXTCOLOR", (i, 0), (i, 0), colors.white)]))

    elements.append(t_tx)
    elements.append(Spacer(1, 24))

    # Firmas
    elements.append(KeepTogether([
        Paragraph("3. Control y Firmas de Responsabilidad", section_title),
        Spacer(1, 20),
        Table([
            [Paragraph("_______________________________<br/>Firma Cajero de Turno", cell_style), Paragraph("_______________________________<br/>Firma Administrador / Auditor", cell_style)]
        ], colWidths=[270, 270])
    ]))

    doc.build(elements)
    stream.seek(0)
    return stream
